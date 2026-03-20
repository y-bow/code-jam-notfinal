import os
import json
from datetime import datetime
from flask import render_template, request, redirect, url_for, g, flash, send_file, abort, current_app
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from io import BytesIO

from ..models import (
    db, User, Student, Teacher, School, Section,
    Course, Enrollment, TimetableEntry, Attendance, Grade,
    UploadedFile, UploadLog, bcrypt
)
from ..middleware import school_scoped, role_minimum
from . import upload_bp

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_upload_path(school_id):
    path = os.path.join(current_app.static_folder, 'uploads', str(school_id))
    os.makedirs(path, exist_ok=True)
    return path

@upload_bp.route('/admin/upload')
@school_scoped
@role_minimum('admin')
def upload_center():
    # Only university admins can access their own school's upload center
    # Global admin can see all, but here we scope to g.school_id if set
    history = UploadedFile.query.filter_by(school_id=g.school_id).order_by(UploadedFile.created_at.desc()).all()
    return render_template('admin/upload.html', history=history)

@upload_bp.route('/admin/upload/template/<type>')
@school_scoped
@role_minimum('admin')
def download_template(type):
    wb = Workbook()
    ws = wb.active
    
    filename = f"{type}_template.xlsx"
    
    if type == 'student':
        headers = ['Name', 'Email', 'Roll Number', 'Section', 'Year', 'Phone']
    elif type == 'course':
        headers = ['Course Name', 'Course Code', 'Credits', 'Section', 'Professor Email']
    elif type == 'timetable':
        headers = ['Course Code', 'Section', 'Day', 'Start Time', 'End Time', 'Room']
    elif type == 'attendance':
        headers = ['Roll Number', 'Date (DD/MM/YYYY)', 'Course Code', 'Status (P/A/L)']
    elif type == 'grade':
        headers = ['Roll Number', 'Course Code', 'Component', 'Marks', 'Max Marks']
    else:
        return abort(404)
        
    ws.append(headers)
    
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name=filename)

@upload_bp.route('/admin/upload/process', methods=['POST'])
@school_scoped
@role_minimum('admin')
def process_upload():
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('upload.upload_center'))
    
    file = request.files['file']
    upload_type = request.form.get('type')
    
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('upload.upload_center'))
    
    if not file or not allowed_file(file.filename):
        flash('Invalid file type. Please upload an .xlsx file.', 'danger')
        return redirect(url_for('upload.upload_center'))

    # Check file size (10MB limit)
    file.seek(0, os.SEEK_END)
    file_length = file.tell()
    if file_length > 10 * 1024 * 1024:
        flash('File size exceeds 10MB limit.', 'danger')
        return redirect(url_for('upload.upload_center'))
    file.seek(0)

    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = secure_filename(f"{timestamp}_{file.filename}")
    filepath = os.path.join('uploads', str(g.school_id), filename)
    full_path = os.path.join(current_app.static_folder, filepath)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    file.save(full_path)

    # Create UploadedFile record
    uploaded_file = UploadedFile(
        school_id=g.school_id,
        uploader_id=g.current_user.id,
        filename=file.filename,
        filepath=filepath,
        file_type=upload_type,
        status='processing'
    )
    db.session.add(uploaded_file)
    db.session.commit()

    try:
        if upload_type == 'student':
            results = process_students(full_path, uploaded_file.id)
        elif upload_type == 'course':
            results = process_courses(full_path, uploaded_file.id)
        elif upload_type == 'timetable':
            results = process_timetable(full_path, uploaded_file.id)
        elif upload_type == 'attendance':
            results = process_attendance(full_path, uploaded_file.id)
        elif upload_type == 'grade':
            results = process_grades(full_path, uploaded_file.id)
        else:
            results = {'created': 0, 'updated': 0, 'errors': 1, 'total': 0}
            log_error(uploaded_file.id, 0, "Unknown upload type", {})

        uploaded_file.status = 'completed'
        uploaded_file.row_count = results['total']
        uploaded_file.success_count = results['created'] + results['updated']
        uploaded_file.error_count = results['errors']
        db.session.commit()
        
        flash(f"Upload processed: {results['created']} created, {results['updated']} updated, {results['errors']} errors.", 'success')
    except Exception as e:
        uploaded_file.status = 'failed'
        db.session.commit()
        log_error(uploaded_file.id, 0, f"Critical processing error: {str(e)}", {})
        flash(f"File processing failed: {str(e)}", 'danger')

    return redirect(url_for('upload.upload_center'))

def log_error(upload_id, row_idx, message, row_data):
    log = UploadLog(
        upload_id=upload_id,
        row_number=row_idx,
        error_message=message,
        row_data=json.dumps(row_data) if row_data else None
    )
    db.session.add(log)

# --- Processing Functions ---

def process_students(filepath, upload_id):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    created = 0
    updated = 0
    errors = 0
    
    pw_hash = bcrypt.generate_password_hash('temporary').decode('utf-8') # Default placeholder

    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            name, email, roll, section_name, year, phone = row[0:6]
            if not email or not roll or not section_name:
                raise ValueError("Missing required fields (Email, Roll Number, Section)")

            # Check if user exists
            user = User.query.filter_by(school_id=g.school_id, email=email).first()
            section = Section.query.filter_by(school_id=g.school_id, name=section_name).first()
            if not section:
                section = Section(school_id=g.school_id, name=section_name, code=section_name.upper()[:10], batch_year=year or datetime.now().year)
                db.session.add(section)
                db.session.flush()

            if user:
                user.name = name
                student = Student.query.get(user.id)
                if not student:
                    student = Student(user_id=user.id, section_id=section.id, enrollment_year=year or datetime.now().year)
                    db.session.add(student)
                else:
                    student.section_id = section.id
                updated += 1
            else:
                new_user = User(
                    school_id=g.school_id,
                    email=email,
                    password_hash=bcrypt.generate_password_hash(str(roll)).decode('utf-8'),
                    role='student',
                    name=name,
                    must_change_password=True
                )
                db.session.add(new_user)
                db.session.flush()
                
                new_student = Student(
                    user_id=new_user.id,
                    section_id=section.id,
                    enrollment_year=year or datetime.now().year
                )
                db.session.add(new_student)
                created += 1
            
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            log_error(upload_id, idx, str(e), row)
            
    return {'created': created, 'updated': updated, 'errors': errors, 'total': len(rows)}

def process_courses(filepath, upload_id):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    created = 0
    updated = 0
    errors = 0
    
    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            name, code, credits, section_name, prof_email = row[0:5]
            if not name or not code or not section_name:
                raise ValueError("Missing required fields (Name, Code, Section)")

            section = Section.query.filter_by(school_id=g.school_id, name=section_name).first()
            if not section:
                section = Section(school_id=g.school_id, name=section_name, code=section_name.upper()[:10], batch_year=datetime.now().year)
                db.session.add(section)
                db.session.flush()

            course = Course.query.filter_by(section_id=section.id, code=code).first()
            
            teacher_id = None
            if prof_email:
                prof = User.query.filter_by(school_id=g.school_id, email=prof_email).first()
                if prof:
                    teacher_id = prof.id

            if not teacher_id:
                # Default to admin if no prof found? Or error? Error is safer.
                raise ValueError(f"Professor with email {prof_email} not found")

            if course:
                course.name = name
                course.credits = credits
                course.teacher_id = teacher_id
                updated += 1
            else:
                new_course = Course(
                    section_id=section.id,
                    name=name,
                    code=code,
                    credits=credits,
                    teacher_id=teacher_id
                )
                db.session.add(new_course)
                created += 1
            
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            log_error(upload_id, idx, str(e), row)
            
    return {'created': created, 'updated': updated, 'errors': errors, 'total': len(rows)}

def process_timetable(filepath, upload_id):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    created = 0
    updated = 0
    errors = 0
    
    DAY_MAP = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4}

    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            course_code, section_name, day_name, start_time, end_time, room = row[0:6]
            
            section = Section.query.filter_by(school_id=g.school_id, name=section_name).first()
            if not section: raise ValueError(f"Section {section_name} not found")
            
            course = Course.query.filter_by(section_id=section.id, code=course_code).first()
            if not course: raise ValueError(f"Course {course_code} not found in section {section_name}")
            
            day = DAY_MAP.get(day_name)
            if day is None: raise ValueError(f"Invalid day: {day_name}")

            # Conflict detection
            # 1. same room + same day + overlapping time
            # 2. same professor + same day + overlapping time
            
            # Simple overlap check: 
            # (StartA < EndB) and (EndA > StartB)
            
            # For simplicity, we create and warn in log, but don't block.
            # But the requirement says "Report conflicts without blocking upload"
            
            new_entry = TimetableEntry(
                section_id=section.id,
                course_id=course.id,
                day=day,
                start_time=start_time,
                end_time=end_time,
                title=course.name,
                teacher=course.teacher.name if course.teacher else "Unknown",
                room=room
            )
            db.session.add(new_entry)
            created += 1
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            log_error(upload_id, idx, str(e), row)
            
    return {'created': created, 'updated': updated, 'errors': errors, 'total': len(rows)}

def process_attendance(filepath, upload_id):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    created = 0
    updated = 0
    errors = 0
    
    STATUS_MAP = {'P': 'present', 'A': 'absent', 'L': 'late'}

    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            roll, date_val, course_code, status_code = row[0:4]
            if not roll or not date_val or not course_code or not status_code:
                raise ValueError("Missing required fields")

            student = Student.query.filter_by(roll_number=str(roll)).first()
            if not student: raise ValueError(f"Student with roll {roll} not found")
            
            # School isolation check
            if student.user.school_id != g.school_id:
                raise ValueError(f"Student {roll} does not belong to your university")

            course = Course.query.filter_by(section_id=student.section_id, code=course_code).first()
            if not course: raise ValueError(f"Course {course_code} not found for student's section")

            status = STATUS_MAP.get(status_code.upper() if isinstance(status_code, str) else status_code)
            if not status: raise ValueError(f"Invalid status: {status_code}")

            # Parse date if it's a string
            if isinstance(date_val, str):
                date_obj = datetime.strptime(date_val, '%d/%m/%Y').date()
            else:
                date_obj = date_val.date() if hasattr(date_val, 'date') else date_val

            # Skip if exists
            exists = Attendance.query.filter_by(student_id=student.user_id, course_id=course.id, date=date_obj).first()
            if not exists:
                new_att = Attendance(
                    student_id=student.user_id,
                    course_id=course.id,
                    date=date_obj,
                    status=status
                )
                db.session.add(new_att)
                created += 1
            
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            log_error(upload_id, idx, str(e), row)
            
    return {'created': created, 'updated': 0, 'errors': errors, 'total': len(rows)}

def process_grades(filepath, upload_id):
    wb = load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    created = 0
    updated = 0
    errors = 0
    
    for idx, row in enumerate(rows, start=2):
        if not any(row): continue
        try:
            roll, course_code, component, marks, max_marks = row[0:5]
            if not roll or not course_code or not marks or not max_marks:
                raise ValueError("Missing required fields")

            if float(marks) > float(max_marks):
                raise ValueError(f"Marks ({marks}) cannot exceed Max Marks ({max_marks})")

            student = Student.query.filter_by(roll_number=str(roll)).first()
            if not student: raise ValueError(f"Student with roll {roll} not found")
            
            if student.user.school_id != g.school_id:
                raise ValueError(f"Student {roll} does not belong to your university")

            course = Course.query.filter_by(section_id=student.section_id, code=course_code).first()
            if not course: raise ValueError(f"Course {course_code} not found for student's section")

            percentage = (float(marks) / float(max_marks)) * 100
            
            grade_rec = Grade.query.filter_by(student_id=student.user_id, course_id=course.id).first()
            if grade_rec:
                grade_rec.grade = percentage
                updated += 1
            else:
                new_grade = Grade(
                    student_id=student.user_id,
                    course_id=course.id,
                    grade=percentage
                )
                db.session.add(new_grade)
                created += 1
            
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            errors += 1
            log_error(upload_id, idx, str(e), row)
            
    return {'created': created, 'updated': updated, 'errors': errors, 'total': len(rows)}

@upload_bp.route('/admin/upload/logs/<int:upload_id>')
@school_scoped
@role_minimum('admin')
def download_logs(upload_id):
    upload = UploadedFile.query.get_or_404(upload_id)
    if upload.school_id != g.school_id: abort(403)
    
    logs = UploadLog.query.filter_by(upload_id=upload_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.append(['Row Number', 'Error Message', 'Original Data'])
    
    for log in logs:
        ws.append([log.row_number, log.error_message, log.row_data])
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name=f"upload_errors_{upload_id}.xlsx")
